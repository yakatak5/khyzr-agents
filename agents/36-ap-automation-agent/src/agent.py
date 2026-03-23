"""
AP Automation Agent
===================
Automates accounts payable workflows: extracts invoice data from PDFs,
matches against purchase orders, flags discrepancies, routes for approval,
and updates the AP ledger.

Built with AWS Strands Agents + Amazon Bedrock AgentCore Runtime (Claude Sonnet).
Deploys as a containerized service on AgentCore — no Lambda required.
"""

import json
import os
import logging
import boto3
from datetime import datetime

# Configure logging early so startup errors appear in CloudWatch
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("ap-automation-agent")
logger.info("AP Automation Agent starting up...")

try:
    from strands import Agent, tool
    from strands.models import BedrockModel
    logger.info("strands-agents imported successfully")
except ImportError as e:
    logger.error(f"Failed to import strands: {e}")
    raise

try:
    from bedrock_agentcore.runtime import BedrockAgentCoreApp
    logger.info("bedrock-agentcore imported successfully")
except ImportError as e:
    logger.error(f"Failed to import bedrock_agentcore: {e}")
    raise


# ---------------------------------------------------------------------------
# AgentCore app
# ---------------------------------------------------------------------------

app = BedrockAgentCoreApp()


# ---------------------------------------------------------------------------
# Tools
# ---------------------------------------------------------------------------

def _fetch_s3_bytes(s3_uri: str):
    """Fetch raw bytes from an S3 URI. Returns (bytes, key) or (None, key)."""
    try:
        parts = s3_uri[5:].split("/", 1)
        bucket, key = parts[0], parts[1]
        s3 = boto3.client("s3")
        obj = s3.get_object(Bucket=bucket, Key=key)
        return obj["Body"].read(), key
    except Exception as e:
        return None, ""


def _parse_excel(raw_bytes: bytes) -> dict:
    """
    Parse an Excel invoice workbook and return a structured invoice dict.

    Supports two common AP invoice layouts:
      Layout A — Single sheet with key:value pairs in columns A:B
                 plus a line-items section starting after a blank row.
      Layout B — First sheet = header info (key:value), second sheet = line items.

    Returns a dict with the same schema as the text/PDF extractor.
    """
    import io
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl not installed — cannot parse Excel files"}

    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
    ws_header = wb.worksheets[0]

    # ── Pass 1: collect all non-empty rows ───────────────────────────────────
    rows = []
    for row in ws_header.iter_rows(values_only=True):
        non_empty = [c for c in row if c is not None]
        if non_empty:
            rows.append([str(c).strip() if c is not None else "" for c in row])

    # ── Pass 2: extract key→value pairs from first two columns ───────────────
    kv = {}
    line_item_start = None
    for i, row in enumerate(rows):
        if len(row) >= 2 and row[0] and row[1]:
            key = row[0].lower().replace(" ", "_").replace(":", "").replace("#", "number")
            kv[key] = row[1]
        # Detect line-items header row
        first = row[0].lower() if row else ""
        if any(h in first for h in ("description", "item", "product", "service")):
            line_item_start = i
            break

    # ── Pass 3: parse line items ─────────────────────────────────────────────
    line_items = []
    if line_item_start is not None:
        # Determine which sheet has line items
        ws_lines = wb.worksheets[1] if len(wb.worksheets) > 1 else ws_header
        header_row = rows[line_item_start]
        # Normalize header names
        col_map = {j: h.lower().replace(" ", "_") for j, h in enumerate(header_row) if h}

        line_rows = list(ws_lines.iter_rows(values_only=True))
        # Skip header row if same sheet
        start = line_item_start + 1 if ws_lines == ws_header else 1
        for row in line_rows[start:]:
            if not any(c for c in row if c is not None):
                continue  # skip blank rows
            item = {}
            for j, cell in enumerate(row):
                col_name = col_map.get(j, f"col_{j}")
                if cell is not None:
                    item[col_name] = cell
            if item:
                line_items.append(item)

    # ── Build canonical invoice dict ─────────────────────────────────────────
    def _find(keys):
        for k in keys:
            for kv_key, val in kv.items():
                if k in kv_key:
                    return val
        return ""

    def _to_float(val):
        if val == "":
            return 0.0
        try:
            return float(str(val).replace(",", "").replace("$", "").strip())
        except Exception:
            return 0.0

    subtotal = _to_float(_find(["subtotal", "net_total", "amount_before"]))
    tax      = _to_float(_find(["tax", "vat", "gst"]))
    total    = _to_float(_find(["total", "amount_due", "grand_total", "invoice_total"]))
    if total == 0.0 and subtotal > 0:
        total = subtotal + tax

    return {
        "invoice_number":        _find(["invoice_number", "invoice_no", "inv_number", "invoice#"]) or "UNKNOWN",
        "vendor_name":           _find(["vendor", "supplier", "from", "company_name"]),
        "vendor_id":             _find(["vendor_id", "supplier_id", "vendor_code"]),
        "invoice_date":          _find(["invoice_date", "date", "issued"]),
        "due_date":              _find(["due_date", "payment_due", "pay_by"]),
        "po_reference":          _find(["po_number", "purchase_order", "po_ref", "po#"]),
        "currency":              _find(["currency", "curr"]) or "USD",
        "subtotal":              subtotal,
        "tax_amount":            tax,
        "total_amount":          total,
        "line_items":            line_items,
        "payment_terms":         _find(["payment_terms", "terms"]),
        "bank_account_last4":    _find(["bank_account", "account_last"]),
        "extraction_method":     "excel_openpyxl",
        "extraction_confidence": 0.92,
        "extracted_at":          datetime.utcnow().isoformat(),
        "source":                "excel",
    }


@tool
def extract_invoice_data(invoice_source: str) -> str:
    """
    Extract structured data from an invoice — supports Excel (.xlsx), PDF, text, or S3 URI.

    Args:
        invoice_source: S3 URI (s3://bucket/key), local file path, or raw text content.
                        Supports: .xlsx, .xls, .pdf, .txt, .csv

    Returns:
        JSON string with extracted invoice fields: vendor, amount, line items, dates, PO reference
    """
    raw_bytes = None
    key = ""

    # ── Fetch from S3 if URI ─────────────────────────────────────────────────
    if invoice_source.startswith("s3://"):
        raw_bytes, key = _fetch_s3_bytes(invoice_source)
        if raw_bytes is None:
            key = invoice_source  # fallback for extension detection

    # ── Local file ───────────────────────────────────────────────────────────
    elif invoice_source.endswith((".xlsx", ".xls", ".pdf", ".txt", ".csv")):
        try:
            with open(invoice_source, "rb") as f:
                raw_bytes = f.read()
            key = invoice_source
        except Exception:
            raw_bytes = None

    # ── Detect file type and route to correct parser ─────────────────────────
    ext = (key or invoice_source).lower().rsplit(".", 1)[-1]

    if ext in ("xlsx", "xls") and raw_bytes is not None:
        invoice_data = _parse_excel(raw_bytes)
        invoice_data["source"] = invoice_source
        return json.dumps(invoice_data, indent=2, default=str)

    # ── Text / plain extraction (existing path) ──────────────────────────────
    raw_text = ""
    if raw_bytes:
        try:
            raw_text = raw_bytes.decode("utf-8")
        except Exception:
            raw_text = ""

    # Realistic mock — mirrors the demo invoice loaded into S3 by Terraform
    invoice_data = {
        "invoice_number": "INV-2024-08821",
        "vendor_name": "Apex Supply Co.",
        "vendor_id": "VND-4492",
        "invoice_date": "2024-03-10",
        "due_date": "2024-04-09",
        "po_reference": "PO-2024-00312",
        "currency": "USD",
        "subtotal": 12450.00,
        "tax_amount": 996.00,
        "total_amount": 13446.00,
        "line_items": [
            {
                "description": "Industrial Filters - 50 units",
                "quantity": 50,
                "unit_price": 189.00,
                "line_total": 9450.00,
            },
            {
                "description": "Maintenance Kit - 10 units",
                "quantity": 10,
                "unit_price": 300.00,
                "line_total": 3000.00,
            },
        ],
        "payment_terms": "Net 30",
        "bank_account_last4": "7823",
        "extraction_confidence": 0.97,
        "extracted_at": datetime.utcnow().isoformat(),
        "source": invoice_source,
        "raw_text_preview": raw_text[:200] if raw_text else "(mock data)",
    }
    return json.dumps(invoice_data, indent=2)


@tool
def match_purchase_order(po_number: str, invoice_data: str) -> str:
    """
    Match an invoice against its referenced purchase order from the ERP system.

    Args:
        po_number: Purchase order number to retrieve and match against
        invoice_data: JSON string of extracted invoice data

    Returns:
        JSON string with match results, variance analysis, and three-way match status
    """
    # In production: queries ERP (SAP, Oracle, NetSuite) via API
    try:
        inv = json.loads(invoice_data)
    except Exception:
        inv = {}

    # Mock PO data — matches the demo invoice exactly for a clean pass
    po_data = {
        "po_number": po_number,
        "vendor_id": "VND-4492",
        "approved_amount": 12450.00,
        "line_items": [
            {"description": "Industrial Filters", "quantity_ordered": 50, "unit_price": 189.00},
            {"description": "Maintenance Kit", "quantity_ordered": 10, "unit_price": 300.00},
        ],
        "status": "approved",
        "approved_by": "procurement@company.com",
        "approved_date": "2024-02-28",
        "goods_receipt_number": "GR-2024-00891",
        "goods_received_date": "2024-03-08",
    }

    inv_total = inv.get("subtotal", 0)
    po_total = po_data["approved_amount"]
    variance = inv_total - po_total
    variance_pct = (variance / po_total * 100) if po_total else 0

    inv_vendor_id = inv.get("vendor_id", "")
    vendor_match = (po_data["vendor_id"] == inv_vendor_id) if inv_vendor_id else True

    match_result = {
        "po_number": po_number,
        "match_status": "matched" if abs(variance_pct) <= 2 and vendor_match else "discrepancy",
        "three_way_match": {
            "po_match": True,
            "receipt_match": True,  # Goods receipt GR-2024-00891 found
            "invoice_match": abs(variance_pct) <= 2,
        },
        "financial_comparison": {
            "po_approved_amount": po_total,
            "invoice_subtotal": inv_total,
            "variance_amount": round(variance, 2),
            "variance_pct": round(variance_pct, 2),
        },
        "vendor_match": vendor_match,
        "vendor_id_on_po": po_data["vendor_id"],
        "vendor_id_on_invoice": inv_vendor_id,
        "goods_receipt": {
            "receipt_number": po_data["goods_receipt_number"],
            "received_date": po_data["goods_received_date"],
        },
        "matched_at": datetime.utcnow().isoformat(),
    }
    return json.dumps(match_result, indent=2)


@tool
def flag_discrepancies(match_result: str) -> str:
    """
    Evaluate match results and flag discrepancies requiring human review.

    Args:
        match_result: JSON string from match_purchase_order

    Returns:
        JSON string listing all discrepancies with severity, codes, and recommended actions
    """
    try:
        data = json.loads(match_result)
    except Exception:
        return json.dumps({"error": "Invalid match result data"})

    discrepancies = []
    fin = data.get("financial_comparison", {})
    variance_pct = abs(fin.get("variance_pct", 0))
    variance_amt = fin.get("variance_amount", 0)

    # Price variance checks
    if variance_pct > 10:
        discrepancies.append({
            "code": "PRICE_VARIANCE_HIGH",
            "severity": "critical",
            "description": f"Invoice amount deviates {variance_pct:.1f}% from PO (${variance_amt:.2f})",
            "action": "Hold payment — escalate to Procurement Manager",
        })
    elif variance_pct > 2:
        discrepancies.append({
            "code": "PRICE_VARIANCE_MINOR",
            "severity": "warning",
            "description": f"Invoice amount deviates {variance_pct:.1f}% from PO",
            "action": "Route to approver with variance note",
        })

    # Vendor mismatch check
    if not data.get("vendor_match", True):
        discrepancies.append({
            "code": "VENDOR_MISMATCH",
            "severity": "critical",
            "description": (
                f"Vendor ID on invoice ({data.get('vendor_id_on_invoice', 'N/A')}) "
                f"does not match PO vendor ({data.get('vendor_id_on_po', 'N/A')})"
            ),
            "action": "Block payment — potential fraud indicator",
        })

    # Goods receipt check
    three_way = data.get("three_way_match", {})
    if not three_way.get("receipt_match", True):
        discrepancies.append({
            "code": "NO_GOODS_RECEIPT",
            "severity": "high",
            "description": "No goods receipt recorded for this PO",
            "action": "Hold payment — confirm delivery with warehouse",
        })

    has_critical = any(d["severity"] == "critical" for d in discrepancies)
    has_high = any(d["severity"] == "high" for d in discrepancies)

    return json.dumps({
        "invoice_status": (
            "hold" if has_critical
            else "review" if (has_high or discrepancies)
            else "approved_for_payment"
        ),
        "discrepancy_count": len(discrepancies),
        "discrepancies": discrepancies,
        "recommended_action": (
            "Block payment and escalate" if has_critical
            else "Hold and review discrepancies" if discrepancies
            else "Clear for payment processing"
        ),
        "evaluated_at": datetime.utcnow().isoformat(),
    }, indent=2)


@tool
def route_for_approval(invoice_number: str, discrepancy_data: str, approver_email: str = "") -> str:
    """
    Route invoice for approval based on discrepancy severity and approval thresholds.

    Args:
        invoice_number: Invoice number to route
        discrepancy_data: JSON string from flag_discrepancies
        approver_email: Override approver email (uses role-based routing if empty)

    Returns:
        JSON string with routing decision, approver assigned, and notification status
    """
    try:
        disc = json.loads(discrepancy_data)
    except Exception:
        disc = {}

    status = disc.get("invoice_status", "review")
    discrepancies = disc.get("discrepancies", [])

    # Role-based routing logic
    if not approver_email:
        has_critical = any(d["severity"] == "critical" for d in discrepancies)
        has_high = any(d["severity"] == "high" for d in discrepancies)
        if has_critical:
            approver_email = os.environ.get("AP_CONTROLLER_EMAIL", "ap-controller@demo.com")
            approver_role = "AP Controller"
            priority = "urgent"
        elif has_high:
            approver_email = os.environ.get("AP_MANAGER_EMAIL", "ap-manager@demo.com")
            approver_role = "AP Manager"
            priority = "high"
        else:
            approver_email = os.environ.get("AP_CLERK_EMAIL", "ap-clerk@demo.com")
            approver_role = "AP Clerk"
            priority = "normal"
    else:
        approver_role = "Manual Override"
        priority = "normal"

    # Simulate sending approval notification via SES/SNS
    # In production: boto3.client('ses').send_email(...)
    routing_result = {
        "invoice_number": invoice_number,
        "routing_status": "routed",
        "approval_required": status != "approved_for_payment",
        "assigned_to": approver_email,
        "approver_role": approver_role,
        "priority": priority,
        "invoice_status": status,
        "discrepancy_summary": (
            f"{len(discrepancies)} discrepancy(ies) found: "
            + ", ".join(d["code"] for d in discrepancies)
            if discrepancies else "No discrepancies — clean invoice"
        ),
        "due_by": datetime.utcnow().strftime("%Y-%m-%d"),
        "notification_sent": True,
        "notification_channel": "email",
        "routed_at": datetime.utcnow().isoformat(),
    }
    return json.dumps(routing_result, indent=2)


@tool
def update_ap_ledger(invoice_number: str, invoice_data: str, approval_status: str = "pending") -> str:
    """
    Update the AP ledger with invoice details and approval status.

    Args:
        invoice_number: Invoice number to record
        invoice_data: JSON string of invoice details
        approval_status: Current status (pending, approved, rejected, paid)

    Returns:
        JSON string with ledger entry confirmation and transaction ID
    """
    try:
        inv = json.loads(invoice_data)
    except Exception:
        inv = {}

    transaction_id = f"AP-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{invoice_number[-5:]}"

    ledger_entry = {
        "transaction_id": transaction_id,
        "invoice_number": invoice_number or inv.get("invoice_number", "UNKNOWN"),
        "vendor_name": inv.get("vendor_name", "Unknown"),
        "vendor_id": inv.get("vendor_id", ""),
        "invoice_date": inv.get("invoice_date", ""),
        "due_date": inv.get("due_date", ""),
        "total_amount": inv.get("total_amount", 0),
        "subtotal": inv.get("subtotal", 0),
        "tax_amount": inv.get("tax_amount", 0),
        "currency": inv.get("currency", "USD"),
        "gl_account": "2000-AP",
        "cost_center": os.environ.get("DEFAULT_COST_CENTER", "CORP-001"),
        "approval_status": approval_status,
        "po_reference": inv.get("po_reference", ""),
        "payment_terms": inv.get("payment_terms", "Net 30"),
        "recorded_at": datetime.utcnow().isoformat(),
        "recorded_by": "AP Automation Agent",
    }

    # Attempt real DynamoDB write
    table_name = os.environ.get("AP_LEDGER_TABLE", "khyzr-ap-automation-demo-ledger")
    region = os.environ.get("AWS_REGION_NAME", os.environ.get("AWS_REGION", "us-east-1"))

    try:
        dynamodb = boto3.resource("dynamodb", region_name=region)
        table = dynamodb.Table(table_name)
        table.put_item(Item=ledger_entry)
        write_status = "written_to_dynamodb"
    except Exception as e:
        write_status = f"simulated (DynamoDB unavailable: {str(e)[:80]})"

    return json.dumps({
        "status": "recorded",
        "transaction_id": transaction_id,
        "write_status": write_status,
        "table": table_name,
        "entry": ledger_entry,
    }, indent=2)


# ---------------------------------------------------------------------------
# Strands Agent definition
# ---------------------------------------------------------------------------

SYSTEM_PROMPT = """You are the AP Automation Agent for Khyzr — an expert accounts payable specialist with deep knowledge of invoice processing, purchase order matching, and financial controls.

Your mission is to automate the end-to-end accounts payable workflow: extract invoice data, perform three-way matching, identify discrepancies, route for approval, and maintain an accurate AP ledger.

When processing an invoice:
1. Extract all structured data from the invoice (vendor, amounts, line items, PO reference, dates)
2. Retrieve the referenced purchase order and perform three-way matching (PO ↔ Invoice ↔ Goods Receipt)
3. Flag any discrepancies with severity classification:
   - **Critical**: Vendor mismatch, >10% price variance, potential duplicate → Block payment
   - **High**: Missing goods receipt, quantity variance → Hold and investigate
   - **Warning**: Minor price variance (2-10%), missing PO reference → Route for review
4. Route invoice to the appropriate approver based on discrepancy severity
5. Update the AP ledger with the invoice record and current approval status

Financial controls you enforce:
- **Segregation of duties**: Extraction, approval, and payment are separate steps
- **Duplicate detection**: Check for duplicate invoice numbers from the same vendor
- **Vendor validation**: Confirm vendor ID matches before processing
- **Payment terms adherence**: Flag invoices approaching due dates for prioritization
- **Early payment discounts**: Identify and flag discount opportunities (e.g., 2/10 Net 30)

Always maintain GAAP compliance and internal audit readiness. Document every decision with clear rationale. Flag 🚨 on any potential fraud indicators (vendor mismatch, unusual bank account changes, round-number amounts). Your work directly impacts cash flow and vendor relationships — be accurate and timely."""

# ---------------------------------------------------------------------------
# Lazy agent initialisation — deferred until first invocation so AgentCore
# runtime startup completes well within the init timeout window.
# ---------------------------------------------------------------------------

_agent = None

def _get_agent():
    global _agent
    if _agent is None:
        logger.info("Initialising BedrockModel + Agent (first invocation)...")
        model = BedrockModel(
            model_id=os.environ.get("BEDROCK_MODEL_ID", "us.anthropic.claude-3-5-sonnet-20241022-v2:0"),
            region_name=os.environ.get("AWS_REGION_NAME", os.environ.get("AWS_REGION", "us-east-1")),
        )
        _agent = Agent(
            model=model,
            tools=[
                extract_invoice_data,
                match_purchase_order,
                flag_discrepancies,
                route_for_approval,
                update_ap_ledger,
            ],
            system_prompt=SYSTEM_PROMPT,
        )
        logger.info("Agent ready.")
    return _agent


# ---------------------------------------------------------------------------
# AgentCore entrypoint
# ---------------------------------------------------------------------------

@app.entrypoint
def invoke(payload):
    """AgentCore entrypoint — receives {"prompt": "..."} """
    user_message = payload.get(
        "prompt",
        payload.get(
            "message",
            "Process demo invoice INV-2024-08821 — extract data, match the PO, "
            "flag discrepancies, route for approval, and update the ledger.",
        ),
    )
    logger.info(f"Received prompt: {user_message[:100]}...")
    try:
        result = _get_agent()(user_message)
        logger.info("Agent invocation completed successfully")
        return {"result": str(result)}
    except Exception as e:
        logger.error(f"Agent invocation failed: {e}", exc_info=True)
        raise


if __name__ == "__main__":
    logger.info("Starting AgentCore HTTP server on port 8080")
    app.run()
